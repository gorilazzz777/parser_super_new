import logging
import traceback

from openpyxl.styles import Alignment, Font, Border, Side, PatternFill

logger = logging.getLogger("app")


class Excel:
    def __init__(self, sheet, fill=None):
        bd = Side(style='thick', color="000000")
        self.sheet = sheet
        self.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
        self.font = Font(bold=True, color='1C1C1C', name='Calibri', size=12)
        self.border = Border(left=bd, top=bd, right=bd, bottom=bd)
        self.fill = PatternFill('solid', fgColor="FFF2CC" if fill is None else fill)

    def border_merge_cells(self, cell, shift):
        for i in range(shift+1):
            self.sheet.cell(row=cell.row, column=cell.column + i).border = self.border

    def set_style(self, cell, shift=0, fill=False):
        cell.alignment = self.alignment
        cell.font = self.font
        if fill:
            cell.fill = self.fill
        self.border_merge_cells(cell, shift=shift)

    def merge_cells(self, cell, shift):
        self.sheet.merge_cells(start_row=cell.row,
                               start_column=cell.column,
                               end_row=cell.row,
                               end_column=cell.column+shift)

    def set_value(self, cell, value, shift=0, fill=False):
        if shift > 0:
            self.merge_cells(cell, shift)
        cell.value = value
        self.set_style(cell=cell, shift=shift, fill=fill)

    def get_value_from_merge_cells(self, cell):
        for column in reversed(range(1, cell.column+1)):
            try:
                if self.sheet.cell(row=cell.row, column=column).value:
                    return self.sheet.cell(row=cell.row, column=column).value
            except:
                logger.error(f'Error in get_value_from_merge_cells: {traceback.format_exc()}')





