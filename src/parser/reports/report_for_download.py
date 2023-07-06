from io import BytesIO

from django.http import HttpResponse
from django.utils.http import urlquote
from openpyxl.workbook import Workbook

from delivery_services.reports.reports import ReportCreator
from delivery_services.services.parsing_tariffs import ManageParsingTariff


class ReportForDownload(ReportCreator):

    def __init__(self, request):
        super().__init__()
        self.request = request


    def _ReportCreator__get_rows(self):
        tariffs, current = ManageParsingTariff().filters(request=self.request)
        return tariffs

    def __fill_headers(self, sheet, packages):
        sheet.cell(row=1, column=1).value = 'Отправитель'
        sheet.cell(row=1, column=2).value = 'Получатель'
        sheet.cell(row=1, column=3).value = 'Регион получателя'
        sheet.cell(row=1, column=4).value = 'Служба доставки'
        sheet.cell(row=1, column=5).value = 'Тариф'
        sheet.cell(row=1, column=6).value = 'Тип доставки'
        sheet.cell(row=1, column=7).value = 'Мин дней'
        sheet.cell(row=1, column=8).value = 'макс дней'
        sheet.cell(row=1, column=8).value = 'Дата обновления'
        for num, package in enumerate(packages):
            sheet.cell(row=1, column=9+num).value = f'{package} кг'

    def _ReportCreator__fill_fields(self, rows):
        wb = Workbook()
        wb.encoding = 'utf-8'
        sheet = wb.active
        self.__fill_headers(sheet=sheet, packages=rows.first().prices.all().order_by('package__weight').values_list('package__weight', flat=True))
        for row, route in enumerate(rows):
            sheet.cell(row=row + 2, column=1).value = route.route.sender_city.name
            sheet.cell(row=row + 2, column=2).value = route.route.receiver_city.name
            sheet.cell(row=row + 2, column=3).value = route.route.receiver_city.region
            sheet.cell(row=row + 2, column=4).value = route.tariff.delivery_servise.name
            sheet.cell(row=row + 2, column=5).value = route.tariff.name
            sheet.cell(row=row + 2, column=6).value = route.tariff.type.name
            sheet.cell(row=row + 2, column=7).value = route.time.max
            sheet.cell(row=row + 2, column=8).value = route.time.min
            for num, cost in enumerate(route.prices.all().order_by('package__weight')):
                sheet.cell(row=row + 2, column=9 + num).value = cost.price
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(output.getvalue(), content_type='application/vnd.ms-excel')
        file_name = f'Отчет.xlsx'
        file_name = urlquote(file_name)
        response['Content-Disposition'] = 'attachment; filename=%s' % file_name
        return response